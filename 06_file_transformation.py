# Author: Duyen Phan

import os
import shutil
import pandas as pd
from openpyxl import load_workbook

# change the address of the base folder according to where you want to store it in your local pc
base_folder = r"C:\Users\a00578308\OneDrive - ONEVIRTUALOFFICE\P&I\APP\CCs_TP_CPD_export\CPD_TP_2025_07"
output_folder = os.path.join(base_folder, "after")
archive_folder = os.path.join(base_folder, "before")
template_path = os.path.join(base_folder, "CPD_CORRECT_TEMPLATE.xlsx")
summary_path = os.path.join(base_folder, "summary.csv")

# create or accept the folders if they are already created
os.makedirs(output_folder, exist_ok=True)
os.makedirs(archive_folder, exist_ok=True)


# check the files every round to make sure these variables are still the same column naming conditions of each type
FUJI_cols = [
    "CUSTOMERCODE", "MATNR", "MAKTX", "GAC", "PGC", "RATE GROUP", "ARTTP",
    "ARP_PRICE", "ARP_CURRENCY", "RATE FACTOR", "TRANSFER_PRICE", "TRANSFER_PRICE_CURRENCY",
    "Installed Base", "Business Catalog", "Brand", "Organization"
]
# FUJI will need a new column "STCCD" at loc = 5

CPD_cols = [
    "CUSTOMERCODE", "MATNR", "MAKTX", "GAC", "PGC", "RATE_GROUP", "ARTTP",
    "ARP_PRICE", "ARP_CURRENCY", "RATE_FACTOR", "ACAB", "TRANSFER_PRICE", "TRANSFER_PRICE_CURRENCY",
    "Installed Base", "Business Catalog", "Brand", "Organization"
]
# CPD will need a new column "STCCD" at loc = 5 + the "ACAB" column will need to be deleted

# keep track of all processed files to avoid a file being processed multiple times in case of duplicates
if os.path.exists(summary_path):
    processed_files = pd.read_csv(summary_path)["file_name"].tolist()
else:
    processed_files = []

summary_entries = []

for file in os.listdir(base_folder):
    if file.endswith(".xlsx") and not file.startswith("CPD_CORRECT_TEMPLATE"):
        file_path = os.path.join(base_folder, file)

        try:
            # skip already processed files:
            if file_path in processed_files:
                raise FileExistsError("FILE ALREADY PROCESSED")

            # read new files and extract the list of columns
            df = pd.read_excel(file_path, engine="openpyxl", dtype=str)
            df.columns = df.columns.str.strip()
            df1 = df.copy()

            # modify columns so FUJI and CPD will have the same format as in the correct template
            if df.columns.tolist() == FUJI_cols:
                df1.insert(loc=5,
                          column="STCCD",
                          value="")
                df_cleaned = df1
            elif df.columns.tolist() == CPD_cols:
                df1.insert(loc=5,
                           column="STCCD",
                           value="")
                df_cleaned = df1.drop(columns="ACAB", axis=1)
                df_cleaned = df_cleaned[df_cleaned["CUSTOMERCODE"].str.contains("Applied filters") == False] # filters depending on the files in the current rounds
            else:
                raise ValueError("WRONG COLUMNS ORDER")

            # copy template and insert new data
            output_path = os.path.join(output_folder, file)
            shutil.copy(template_path, output_path)

            wb = load_workbook(output_path)
            ws = wb["Export"]
            ws.delete_rows(2, ws.max_row) # make sure the output sheet is empty except the header row
            for row in df_cleaned.itertuples(index=False, name=None):
                ws.append(row)
            wb.save(output_path)

            # move the processed file to the archive folder:
            before_path = os.path.join(archive_folder, file)
            if os.path.exists(before_path):
                raise FileExistsError("FILE ALREADY PROCESSED") # skip files already inside the archive folder
            shutil.move(file_path, before_path)

            # summary counts:
            row_count = len(df_cleaned)
            summary_entries.append({"file_name": file, "row_count": row_count})

            print(f"{file}: PROCESSED SUCCESSFULLY")

        except Exception as e:
            print(f"Error processing {file}: {e}")

# append to or create summary.csv
if summary_entries: # not empty
    df_summary = pd.DataFrame(data=summary_entries, index=None)
    if os.path.exists(summary_path):
        df_summary.to_csv(summary_path, header=False, index=False, mode="a")
    else:
        df_summary.to_csv(summary_path, header=True, index=False)








